from __future__ import annotations

import json
import importlib.util
import os
import subprocess
import sys
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook


def _json_default(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def export_json(data: Dict[str, Any]) -> bytes:
    return json.dumps(
        data,
        ensure_ascii=False,
        indent=2,
        default=_json_default,
    ).encode("utf-8")


def export_markdown(content: str) -> bytes:
    return content.encode("utf-8")


def export_xlsx(data: Dict[str, Any]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "经营摘要"
    summary.append(["指标", "数值"])
    for key, value in data.items():
        if isinstance(value, (str, int, float)) or value is None:
            summary.append([key, value])
    actions = workbook.create_sheet("管理动作")
    actions.append(["优先级", "管理动作", "责任部门"])
    for item in data.get("top_actions", []):
        actions.append([item.get("priority"), item.get("action"), item.get("owner")])
    risks = workbook.create_sheet("高风险订单")
    risks.append(["订单", "风险分", "风险等级", "影响金额"])
    for item in data.get("high_risk_orders", []):
        risks.append([
            item.get("sales_order_code"),
            item.get("risk_score"),
            item.get("risk_level"),
            item.get("potential_amount"),
        ])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(width, 42)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_docx(title: str, lines: Iterable[str]) -> bytes:
    paragraphs = "".join(
        "<w:p><w:r><w:t xml:space=\"preserve\">"
        f"{escape(line)}"
        "</w:t></w:r></w:p>"
        for line in [title, *list(lines)]
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{paragraphs}<w:sectPr/></w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    relationships = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/></Relationships>'
    )
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", relationships)
        archive.writestr("word/document.xml", document)
    return buffer.getvalue()


def export_pdf(title: str, lines: List[str]) -> bytes:
    root = Path(__file__).resolve().parents[1]
    renderer = root / "scripts" / "render_report_pdf.py"
    configured_python = os.environ.get("REPORT_PDF_PYTHON")
    bundled_python = Path(
        "C:/Users/25301/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/python.exe"
    )
    if configured_python:
        python_executable = configured_python
    elif importlib.util.find_spec("reportlab") is not None:
        python_executable = sys.executable
    elif bundled_python.exists():
        python_executable = str(bundled_python)
    else:
        raise RuntimeError("PDF导出缺少ReportLab，请安装requirements.txt中的依赖")
    payload = json.dumps({"title": title, "lines": lines}, ensure_ascii=False).encode("utf-8")
    completed = subprocess.run(
        [python_executable, str(renderer)],
        input=payload,
        capture_output=True,
        check=False,
        cwd=str(root),
    )
    if completed.returncode != 0 or not completed.stdout.startswith(b"%PDF"):
        details = completed.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(f"PDF生成失败: {details or '渲染器未返回有效PDF'}")
    return completed.stdout
