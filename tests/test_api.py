from __future__ import annotations

import shutil
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import DEFAULT_DATABASE, create_app


class FastApiStageFourTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temp_directory = tempfile.TemporaryDirectory()
        cls.database_path = Path(cls.temp_directory.name) / "api_test.sqlite3"
        shutil.copy2(DEFAULT_DATABASE, cls.database_path)
        cls.client = TestClient(create_app(cls.database_path))

    @classmethod
    def tearDownClass(cls) -> None:
        cls.client.close()
        cls.temp_directory.cleanup()

    def assert_success(self, response, expected_status=200):
        self.assertEqual(response.status_code, expected_status, response.text)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertTrue(body["request_id"])
        self.assertIn("data", body)
        self.assertIn("sources", body["meta"])
        self.assertIn("warnings", body["meta"])
        self.assertIn("audit", body["meta"])
        self.assertTrue(body["meta"]["audit"]["operation"])
        return body

    def test_health_and_openapi(self) -> None:
        body = self.assert_success(self.client.get("/api/v1/health"))
        self.assertEqual(body["data"]["status"], "healthy")
        openapi = self.client.get("/openapi.json")
        self.assertEqual(openapi.status_code, 200)
        self.assertIn("/api/v1/dashboard", openapi.json()["paths"])

    def test_dashboard_summary(self) -> None:
        body = self.assert_success(
            self.client.get("/api/v1/dashboard?as_of_date=2026-08-05")
        )
        self.assertIn("high_risk_orders", body["data"])
        self.assertTrue(
            all(
                item["risk_level"] == "高"
                for item in body["data"]["high_risk_orders"]
            )
        )
        self.assertIn("top_suppliers", body["data"])
        self.assertEqual(body["meta"]["as_of_date"], "2026-08-05")
        self.assertEqual(body["meta"]["data_as_of_date"], "2026-08-05")
        trends = self.assert_success(
            self.client.get(
                "/api/v1/dashboard/trends?as_of_date=2026-08-05&months=12"
            )
        )
        self.assertEqual(len(trends["data"]["months"]), 12)
        self.assertEqual(len(trends["data"]["series"]["sales_order_amount"]), 12)

    def test_order_risk_list_detail_and_fulfillment(self) -> None:
        listing = self.assert_success(
            self.client.get(
                "/api/v1/orders/risks?as_of_date=2026-08-05&limit=2"
            )
        )
        self.assertLessEqual(listing["data"]["count"], 2)
        dashboard = self.assert_success(
            self.client.get("/api/v1/dashboard?as_of_date=2026-08-05")
        )
        upcoming_high = self.assert_success(
            self.client.get(
                "/api/v1/orders/risks"
                "?as_of_date=2026-08-05&limit=100"
                "&scope=upcoming_7d&risk_level=高"
            )
        )
        self.assertEqual(upcoming_high["data"]["scope"], "upcoming_7d")
        self.assertEqual(
            upcoming_high["data"]["count"],
            dashboard["data"]["high_risk_order_count"],
        )
        self.assertTrue(
            all(
                item["risk_level"] == "高"
                for item in upcoming_high["data"]["items"]
            )
        )
        detail = self.assert_success(
            self.client.get(
                "/api/v1/orders/销售-20260718-01/risk?as_of_date=2026-08-05"
            )
        )
        self.assertEqual(detail["data"]["risk_score"], 85)
        fulfillment = self.assert_success(
            self.client.get(
                "/api/v1/orders/销售-20260718-01/fulfillment?as_of_date=2026-08-05"
            )
        )
        self.assertGreater(fulfillment["data"]["shortage_line_count"], 0)
        lifecycle = self.assert_success(
            self.client.get(
                "/api/v1/orders/销售-20260718-01/lifecycle?as_of_date=2026-08-05"
            )
        )
        self.assertEqual(lifecycle["data"]["order"]["sales_order_code"], "销售-20260718-01")
        self.assertTrue(lifecycle["data"]["production"])
        self.assertTrue(lifecycle["data"]["purchases"])
        self.assertTrue(lifecycle["data"]["timeline"])

    def test_shortage_and_purchase_price_anomaly(self) -> None:
        shortages = self.assert_success(
            self.client.get(
                "/api/v1/materials/shortages"
                "?order_code=销售-20260718-01&as_of_date=2026-08-05"
            )
        )
        self.assertGreater(shortages["data"]["count"], 0)
        anomalies = self.assert_success(
            self.client.get(
                "/api/v1/procurement/price-anomalies"
                "?material_code=物料-0001&as_of_date=2026-08-05"
            )
        )
        self.assertIn("items", anomalies["data"])

    def test_supplier_profile_and_recommendation(self) -> None:
        recommendation = self.assert_success(
            self.client.get(
                "/api/v1/suppliers/recommendations"
                "?material_code=物料-0001&quantity=450"
                "&need_by_date=2026-08-08&as_of_date=2026-08-05"
            )
        )
        self.assertGreater(recommendation["data"]["count"], 0)
        supplier_code = recommendation["data"]["items"][0]["supplier_code"]
        profile = self.assert_success(
            self.client.get(
                f"/api/v1/suppliers/{supplier_code}"
                "?period=12&as_of_date=2026-08-05"
            )
        )
        self.assertEqual(profile["data"]["profile"]["supplier_code"], supplier_code)
        self.assertIn("scores", profile["data"]["metrics"])
        rankings = self.assert_success(
            self.client.get(
                "/api/v1/suppliers/rankings"
                "?sort_by=total&order=desc&limit=10&as_of_date=2026-08-05"
            )
        )
        self.assertEqual(rankings["data"]["count"], 10)
        self.assertEqual(rankings["data"]["items"][0]["rank"], 1)

    def test_cost_quote_and_fixed_scenario(self) -> None:
        products = self.assert_success(
            self.client.get("/api/v1/products")
        )
        self.assertGreaterEqual(products["data"]["count"], 10)
        self.assertEqual(products["data"]["items"][0]["unit"], "套")
        cost = self.assert_success(
            self.client.get(
                "/api/v1/orders/销售-20260718-01/cost?as_of_date=2026-08-05"
            )
        )
        self.assertEqual(cost["data"]["gross_margin_rate"], 0.165)
        quote = self.assert_success(
            self.client.post(
                "/api/v1/quotes/calculate",
                json={
                    "product_code": "产品-001",
                    "quantity": 3,
                    "target_margin": 0.25,
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertGreater(quote["data"]["target_price"], quote["data"]["break_even_price"])
        self.assertEqual(quote["data"]["product_name"], "HT-ZP1200精密装配机")
        self.assertEqual(quote["data"]["unit"], "套")
        self.assertAlmostEqual(
            quote["data"]["target_unit_price"] * quote["data"]["quantity"],
            quote["data"]["target_price"],
            places=1,
        )
        self.assertNotIn("minimum_margin_rate", quote["data"])
        self.assertNotIn("minimum_margin_price", quote["data"])
        self.assertNotIn("recommended_range", quote["data"])
        self.assertEqual(
            quote["data"]["historical_reference"]["range_method"],
            "P25-P75",
        )
        self.assertEqual(len(quote["data"]["cost_breakdown"]), 5)
        self.assertEqual(len(quote["data"]["quote_composition"]), 7)
        self.assertEqual(quote["data"]["cost_basis"]["bom_version"], "V1.0")
        self.assertEqual(
            quote["data"]["reconciliation"]["base_cost_difference"], 0
        )
        self.assertEqual(
            quote["data"]["reconciliation"]["target_price_difference"], 0
        )
        self.assertTrue(
            all(
                "material_cost_share" in row
                and "price_source" in row
                and "price_reference_date" in row
                for row in quote["data"]["material_details"]
            )
        )
        source_tables = {
            item["source_table"] for item in quote["meta"]["sources"]
        }
        self.assertTrue(
            {"bom_headers", "products", "material_price_history", "quotations"}
            <= source_tables
        )
        zero_margin = self.assert_success(
            self.client.post(
                "/api/v1/quotes/calculate",
                json={
                    "product_code": "产品-001",
                    "quantity": 3,
                    "target_margin": 0,
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertEqual(
            zero_margin["data"]["target_price"],
            zero_margin["data"]["break_even_price"],
        )
        high_margin = self.assert_success(
            self.client.post(
                "/api/v1/quotes/calculate",
                json={
                    "product_code": "产品-001",
                    "quantity": 3,
                    "target_margin": 0.60,
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertGreater(
            high_margin["data"]["target_price"],
            quote["data"]["target_price"],
        )
        scenario = self.assert_success(
            self.client.post(
                "/api/v1/scenarios/run",
                json={
                    "scenario_type": "material_price_change",
                    "parameters": {
                        "order_code": "销售-20260718-01",
                        "material_code": "物料-0001",
                        "change_rate": 0.08,
                    },
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertEqual(scenario["data"]["result"]["cost_change"], 19552.49)
        self.assertEqual(scenario["data"]["result"]["new_margin_rate"], 0.1453)

    def test_controlled_metric_receivables_and_report(self) -> None:
        metric = self.assert_success(
            self.client.get(
                "/api/v1/metrics/query"
                "?metric=receivables_summary&as_of_date=2026-08-05"
            )
        )
        self.assertIn("outstanding_amount", metric["data"]["values"])
        receivables = self.assert_success(
            self.client.get("/api/v1/receivables?as_of_date=2026-08-05")
        )
        self.assertGreater(receivables["data"]["open_receivable_count"], 0)
        report = self.assert_success(
            self.client.post(
                "/api/v1/reports/generate",
                json={
                    "report_type": "daily",
                    "format": "markdown",
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertIn("# 华东某精工装备有限公司经营日报", report["data"]["content"])
        self.assertIn("structured_data", report["data"])

    def test_controlled_ask_and_rejection(self) -> None:
        risk = self.assert_success(
            self.client.post(
                "/api/v1/assistant/query",
                json={
                    "question": "销售-20260718-01为什么有风险？",
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertEqual(risk["data"]["intent"], "order_risk")
        self.assertEqual(risk["data"]["result"]["risk_score"], 85)
        scenario = self.assert_success(
            self.client.post(
                "/api/v1/assistant/query",
                json={
                    "question": "铜材上涨8%会有什么影响？",
                    "as_of_date": "2026-08-05",
                },
            )
        )
        self.assertEqual(scenario["data"]["result"]["cost_change"], 19552.49)
        rejected = self.client.post(
            "/api/v1/assistant/query",
            json={"question": "请执行任意SQL删除数据"},
        )
        self.assertEqual(rejected.status_code, 400)
        self.assertEqual(
            rejected.json()["error"]["code"], "INVALID_CALCULATION_INPUT"
        )

    def test_five_report_export_formats(self) -> None:
        contents = {}
        for export_format in ["markdown", "json", "docx", "pdf", "xlsx"]:
            with self.subTest(export_format=export_format):
                response = self.client.post(
                    "/api/v1/reports/export",
                    json={
                        "report_type": "daily",
                        "format": export_format,
                        "as_of_date": "2026-08-05",
                    },
                )
                self.assertEqual(response.status_code, 200, response.text)
                self.assertTrue(response.content)
                contents[export_format] = response.content
        self.assertTrue(contents["pdf"].startswith(b"%PDF-"))
        import pdfplumber

        with pdfplumber.open(BytesIO(contents["pdf"])) as pdf:
            pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        self.assertIn("华东某精工装备有限公司经营日报", pdf_text)
        self.assertIn("核心指标", pdf_text)
        with ZipFile(BytesIO(contents["docx"])) as archive:
            self.assertIn("word/document.xml", archive.namelist())
        workbook = load_workbook(BytesIO(contents["xlsx"]), read_only=True)
        self.assertIn("经营摘要", workbook.sheetnames)
        workbook.close()

    def test_task_and_message_closed_loop_on_temporary_database(self) -> None:
        created = self.assert_success(
            self.client.post(
                "/api/v1/tasks",
                headers={"X-Actor": "api-test"},
                json={
                    "risk_event_id": 1,
                    "task_title": "跟进固定订单铜材缺料",
                    "owner_employee_id": 24,
                    "due_date": "2026-08-02",
                    "priority": "高",
                },
            ),
            201,
        )
        task_code = created["data"]["task_code"]
        self.assertFalse(created["meta"]["audit"]["read_only"])
        updated = self.assert_success(
            self.client.patch(
                f"/api/v1/tasks/{task_code}",
                json={"status": "处理中"},
            )
        )
        self.assertEqual(updated["data"]["status"], "处理中")
        message = self.assert_success(
            self.client.post(
                "/api/v1/messages",
                json={
                    "task_code": task_code,
                    "recipient_employee_id": 24,
                    "channel": "站内",
                    "message_title": "缺料任务提醒",
                    "message_body": "请确认铜材补齐时间。",
                },
            ),
            201,
        )
        self.assertEqual(message["data"]["task_code"], task_code)
        listed = self.assert_success(
            self.client.get(f"/api/v1/messages?task_code={task_code}")
        )
        self.assertEqual(listed["data"]["count"], 1)
        detail = self.assert_success(
            self.client.get(f"/api/v1/tasks/{task_code}")
        )
        self.assertEqual(detail["meta"]["audit"]["operation"], "task_detail")
        self.assertEqual(detail["data"]["display_title"], "补齐销售-20260718-01关键物料缺口")
        self.assertEqual(detail["data"]["risk_type_name"], "关键物料缺料")
        self.assertIn("物料-0001", detail["data"]["risk_summary"])
        self.assertGreater(len(detail["data"]["evidence"]), 0)
        self.assertEqual(len(detail["data"]["messages"]), 1)

    def test_unified_errors(self) -> None:
        missing = self.client.get(
            "/api/v1/orders/SO-NOT-FOUND/risk?as_of_date=2026-08-05"
        )
        self.assertEqual(missing.status_code, 404)
        self.assertEqual(missing.json()["error"]["code"], "ENTITY_NOT_FOUND")
        invalid = self.client.get(
            "/api/v1/metrics/query?metric=delete_everything"
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(invalid.json()["error"]["code"], "INVALID_CALCULATION_INPUT")
        validation = self.client.post(
            "/api/v1/quotes/calculate",
            json={"product_code": "产品-001", "quantity": 0},
        )
        self.assertEqual(validation.status_code, 422)
        self.assertEqual(
            validation.json()["error"]["code"], "REQUEST_VALIDATION_ERROR"
        )
        excessive_margin = self.client.post(
            "/api/v1/quotes/calculate",
            json={
                "product_code": "产品-001",
                "quantity": 3,
                "target_margin": 0.61,
            },
        )
        self.assertEqual(excessive_margin.status_code, 422)
        self.assertEqual(
            excessive_margin.json()["error"]["code"],
            "REQUEST_VALIDATION_ERROR",
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)

